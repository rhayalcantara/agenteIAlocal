"""
Excel Tool — Herramientas para manipular archivos Excel con openpyxl.

Operaciones disponibles:
  Lectura    : info, leer, leer_formulas, listar_hojas, buscar
  Escritura  : escribir, crear, reemplazar
  Hojas      : crear_hoja, eliminar_hoja, renombrar_hoja, copiar_hoja, mover_hoja
  Estructura : insertar_filas, eliminar_filas, insertar_columnas, eliminar_columnas, ordenar
  Formato    : formato_celdas, formato_condicional, auto_ajustar_columnas
  Objetos    : crear_tabla, agregar_grafico, proteger_hoja, hipervinculo
"""
import os
import re
from logger import get_logger

logger = get_logger("excel_tool")


# ── Helpers internos ──────────────────────────────────────────────────────────

def _importar_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")


def _cargar(ruta: str, data_only: bool = False):
    ox = _importar_openpyxl()
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta}")
    return ox.load_workbook(ruta, data_only=data_only)


def _hoja(wb, nombre: str = None):
    if nombre:
        if nombre not in wb.sheetnames:
            raise ValueError(f"Hoja '{nombre}' no encontrada. Disponibles: {wb.sheetnames}")
        return wb[nombre]
    return wb.active


# ── INFO ──────────────────────────────────────────────────────────────────────

def info(ruta: str) -> str:
    """Información general: hojas, dimensiones, propiedades."""
    try:
        wb = _cargar(ruta, data_only=True)
        lineas = [f"📊 {os.path.basename(ruta)}",
                  f"Hojas ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}"]
        for nombre in wb.sheetnames:
            ws = wb[nombre]
            lineas.append(f"  • {nombre}: {ws.max_row} filas × {ws.max_column} columnas")
        return "\n".join(lineas)
    except Exception as e:
        logger.error(f"excel.info: {e}")
        return f"Error: {e}"


# ── LEER ──────────────────────────────────────────────────────────────────────

def leer(ruta: str, hoja: str = None, rango: str = None, max_filas: int = 100) -> str:
    """Lee datos de un rango y retorna tabla formateada en texto."""
    try:
        wb = _cargar(ruta, data_only=True)
        ws = _hoja(wb, hoja)
        filas_raw = list(ws[rango]) if rango else list(ws.iter_rows())
        filas_raw = filas_raw[:max_filas]
        if not filas_raw:
            return "(sin datos)"
        tabla = [[str(c.value if c.value is not None else "") for c in f] for f in filas_raw]
        anchos = [max(len(tabla[r][c]) for r in range(len(tabla))) for c in range(len(tabla[0]))]
        return "\n".join(" | ".join(v.ljust(anchos[i]) for i, v in enumerate(f)) for f in tabla)
    except Exception as e:
        logger.error(f"excel.leer: {e}")
        return f"Error: {e}"


def leer_formulas(ruta: str, hoja: str = None, rango: str = None) -> str:
    """Lee las fórmulas (no valores calculados) de un rango."""
    try:
        wb = _cargar(ruta, data_only=False)
        ws = _hoja(wb, hoja)
        filas = list(ws[rango]) if rango else list(ws.iter_rows())
        resultado = [
            f"{c.coordinate}: {c.value}"
            for f in filas for c in f
            if c.value and str(c.value).startswith("=")
        ]
        return "\n".join(resultado) if resultado else "(sin fórmulas en el rango)"
    except Exception as e:
        logger.error(f"excel.leer_formulas: {e}")
        return f"Error: {e}"


def listar_hojas(ruta: str) -> str:
    """Lista todas las hojas del archivo."""
    try:
        wb = _cargar(ruta)
        return f"Hojas ({len(wb.sheetnames)}): " + ", ".join(f"'{h}'" for h in wb.sheetnames)
    except Exception as e:
        logger.error(f"excel.listar_hojas: {e}")
        return f"Error: {e}"


def buscar(ruta: str, texto: str, hoja: str = None) -> str:
    """Busca texto en celdas. Si hoja=None busca en todas las hojas."""
    try:
        wb = _cargar(ruta, data_only=True)
        hojas = [wb[hoja]] if (hoja and hoja in wb.sheetnames) else [wb[h] for h in wb.sheetnames]
        resultados = [
            f"[{ws.title}] {c.coordinate}: {c.value}"
            for ws in hojas
            for f in ws.iter_rows()
            for c in f
            if c.value and texto.lower() in str(c.value).lower()
        ]
        if not resultados:
            return f"No se encontró '{texto}'"
        return f"Encontradas {len(resultados)} ocurrencias:\n" + "\n".join(resultados[:50])
    except Exception as e:
        logger.error(f"excel.buscar: {e}")
        return f"Error: {e}"


# ── ESCRIBIR ──────────────────────────────────────────────────────────────────

def escribir(ruta: str, datos: list, hoja: str = None, celda_inicio: str = "A1") -> str:
    """Escribe lista de listas en la hoja desde celda_inicio.

    datos: [["A", "B"], [1, 2], ...]
    """
    try:
        ox = _importar_openpyxl()
        wb = _cargar(ruta) if os.path.exists(ruta) else ox.Workbook()
        ws = _hoja(wb, hoja)

        m = re.match(r"([A-Za-z]+)(\d+)", celda_inicio)
        from openpyxl.utils import column_index_from_string
        col_ini = column_index_from_string(m.group(1)) if m else 1
        row_ini = int(m.group(2)) if m else 1

        for r, fila in enumerate(datos):
            for c, valor in enumerate(fila):
                ws.cell(row=row_ini + r, column=col_ini + c, value=valor)

        wb.save(ruta)
        return f"✅ {len(datos)} filas × {max((len(f) for f in datos), default=0)} cols escritas desde {celda_inicio} en '{ws.title}'"
    except Exception as e:
        logger.error(f"excel.escribir: {e}")
        return f"Error: {e}"


def crear(ruta: str, hojas: list = None) -> str:
    """Crea un nuevo archivo Excel vacío con las hojas indicadas."""
    try:
        ox = _importar_openpyxl()
        wb = ox.Workbook()
        if hojas:
            wb.active.title = hojas[0]
            for nombre in hojas[1:]:
                wb.create_sheet(nombre)
        else:
            wb.active.title = "Hoja1"
        os.makedirs(os.path.dirname(os.path.abspath(ruta)), exist_ok=True)
        wb.save(ruta)
        return f"✅ Archivo creado: {ruta} | Hojas: {wb.sheetnames}"
    except Exception as e:
        logger.error(f"excel.crear: {e}")
        return f"Error: {e}"


def reemplazar(ruta: str, texto_buscar: str, texto_reemplazar: str, hoja: str = None) -> str:
    """Reemplaza texto en celdas de texto."""
    try:
        wb = _cargar(ruta)
        hojas = [wb[hoja]] if (hoja and hoja in wb.sheetnames) else [wb[h] for h in wb.sheetnames]
        count = 0
        for ws in hojas:
            for f in ws.iter_rows():
                for c in f:
                    if isinstance(c.value, str) and texto_buscar in c.value:
                        c.value = c.value.replace(texto_buscar, texto_reemplazar)
                        count += 1
        wb.save(ruta)
        return f"✅ {count} ocurrencias reemplazadas: '{texto_buscar}' → '{texto_reemplazar}'"
    except Exception as e:
        logger.error(f"excel.reemplazar: {e}")
        return f"Error: {e}"


# ── GESTIÓN DE HOJAS ──────────────────────────────────────────────────────────

def crear_hoja(ruta: str, nombre: str, posicion: int = None) -> str:
    try:
        wb = _cargar(ruta)
        if nombre in wb.sheetnames:
            return f"La hoja '{nombre}' ya existe."
        wb.create_sheet(nombre, index=posicion)
        wb.save(ruta)
        return f"✅ Hoja '{nombre}' creada. Hojas: {wb.sheetnames}"
    except Exception as e:
        logger.error(f"excel.crear_hoja: {e}")
        return f"Error: {e}"


def eliminar_hoja(ruta: str, nombre: str) -> str:
    try:
        wb = _cargar(ruta)
        if nombre not in wb.sheetnames:
            return f"Hoja '{nombre}' no encontrada. Disponibles: {wb.sheetnames}"
        if len(wb.sheetnames) == 1:
            return "No se puede eliminar la única hoja del archivo."
        del wb[nombre]
        wb.save(ruta)
        return f"✅ Hoja '{nombre}' eliminada. Restantes: {wb.sheetnames}"
    except Exception as e:
        logger.error(f"excel.eliminar_hoja: {e}")
        return f"Error: {e}"


def renombrar_hoja(ruta: str, nombre_actual: str, nombre_nuevo: str) -> str:
    try:
        wb = _cargar(ruta)
        ws = _hoja(wb, nombre_actual)
        ws.title = nombre_nuevo
        wb.save(ruta)
        return f"✅ Hoja renombrada: '{nombre_actual}' → '{nombre_nuevo}'"
    except Exception as e:
        logger.error(f"excel.renombrar_hoja: {e}")
        return f"Error: {e}"


def copiar_hoja(ruta: str, origen: str, destino: str) -> str:
    try:
        wb = _cargar(ruta)
        if origen not in wb.sheetnames:
            return f"Hoja '{origen}' no encontrada."
        copia = wb.copy_worksheet(wb[origen])
        copia.title = destino
        wb.save(ruta)
        return f"✅ Hoja '{origen}' copiada como '{destino}'"
    except Exception as e:
        logger.error(f"excel.copiar_hoja: {e}")
        return f"Error: {e}"


def mover_hoja(ruta: str, nombre: str, posicion: int) -> str:
    try:
        wb = _cargar(ruta)
        if nombre not in wb.sheetnames:
            return f"Hoja '{nombre}' no encontrada."
        actual = wb.sheetnames.index(nombre)
        wb.move_sheet(nombre, offset=posicion - actual)
        wb.save(ruta)
        return f"✅ Hoja '{nombre}' movida a posición {posicion}. Orden: {wb.sheetnames}"
    except Exception as e:
        logger.error(f"excel.mover_hoja: {e}")
        return f"Error: {e}"


# ── ESTRUCTURA (FILAS/COLUMNAS) ───────────────────────────────────────────────

def insertar_filas(ruta: str, hoja: str, fila: int, cantidad: int = 1) -> str:
    try:
        wb = _cargar(ruta)
        _hoja(wb, hoja).insert_rows(fila, cantidad)
        wb.save(ruta)
        return f"✅ {cantidad} fila(s) insertada(s) en posición {fila}"
    except Exception as e:
        return f"Error: {e}"


def eliminar_filas(ruta: str, hoja: str, fila: int, cantidad: int = 1) -> str:
    try:
        wb = _cargar(ruta)
        _hoja(wb, hoja).delete_rows(fila, cantidad)
        wb.save(ruta)
        return f"✅ {cantidad} fila(s) eliminada(s) desde posición {fila}"
    except Exception as e:
        return f"Error: {e}"


def insertar_columnas(ruta: str, hoja: str, columna: int, cantidad: int = 1) -> str:
    try:
        wb = _cargar(ruta)
        _hoja(wb, hoja).insert_cols(columna, cantidad)
        wb.save(ruta)
        return f"✅ {cantidad} columna(s) insertada(s) en posición {columna}"
    except Exception as e:
        return f"Error: {e}"


def eliminar_columnas(ruta: str, hoja: str, columna: int, cantidad: int = 1) -> str:
    try:
        wb = _cargar(ruta)
        _hoja(wb, hoja).delete_cols(columna, cantidad)
        wb.save(ruta)
        return f"✅ {cantidad} columna(s) eliminada(s) desde posición {columna}"
    except Exception as e:
        return f"Error: {e}"


def ordenar(ruta: str, hoja: str, columna: int,
            ascendente: bool = True, fila_inicio: int = 2) -> str:
    """Ordena los datos por columna (índice 1-based). fila_inicio=2 omite encabezado."""
    try:
        wb = _cargar(ruta, data_only=True)
        ws = _hoja(wb, hoja)
        filas = list(ws.iter_rows(min_row=fila_inicio, values_only=True))
        col_idx = columna - 1
        filas.sort(key=lambda f: (f[col_idx] is None, f[col_idx]), reverse=not ascendente)

        wb2 = _cargar(ruta)
        ws2 = _hoja(wb2, ws.title)
        for r_idx, fila in enumerate(filas):
            for c_idx, valor in enumerate(fila):
                ws2.cell(row=fila_inicio + r_idx, column=c_idx + 1, value=valor)
        wb2.save(ruta)
        return f"✅ Datos ordenados por columna {columna} ({'asc' if ascendente else 'desc'})"
    except Exception as e:
        logger.error(f"excel.ordenar: {e}")
        return f"Error: {e}"


# ── FORMATO ───────────────────────────────────────────────────────────────────

def formato_celdas(ruta: str, hoja: str, rango: str,
                   negrita: bool = None, cursiva: bool = None,
                   color_fondo: str = None, color_texto: str = None,
                   tamanio_fuente: int = None, alineacion: str = None) -> str:
    """Aplica formato a celdas.

    color_fondo / color_texto: hex sin '#' (ej: 'FFFF00' amarillo)
    alineacion: 'left' | 'center' | 'right'
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = _cargar(ruta)
        ws = _hoja(wb, hoja)
        for fila in ws[rango]:
            for c in fila:
                font_kw = {}
                if negrita is not None: font_kw["bold"] = negrita
                if cursiva is not None: font_kw["italic"] = cursiva
                if tamanio_fuente: font_kw["size"] = tamanio_fuente
                if color_texto: font_kw["color"] = color_texto.lstrip("#").upper()
                if font_kw:
                    c.font = Font(**font_kw)
                if color_fondo:
                    c.fill = PatternFill(fill_type="solid",
                                         fgColor=color_fondo.lstrip("#").upper())
                if alineacion:
                    c.alignment = Alignment(horizontal=alineacion)
        wb.save(ruta)
        return f"✅ Formato aplicado al rango {rango}"
    except Exception as e:
        logger.error(f"excel.formato_celdas: {e}")
        return f"Error: {e}"


def auto_ajustar_columnas(ruta: str, hoja: str = None) -> str:
    """Auto-ajusta el ancho de todas las columnas al contenido."""
    try:
        wb = _cargar(ruta)
        ws = _hoja(wb, hoja)
        for col in ws.columns:
            letra = col[0].column_letter
            ancho = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[letra].width = min(ancho + 2, 60)
        wb.save(ruta)
        return f"✅ Columnas auto-ajustadas en '{ws.title}'"
    except Exception as e:
        logger.error(f"excel.auto_ajustar_columnas: {e}")
        return f"Error: {e}"


def formato_condicional(ruta: str, hoja: str, rango: str,
                        tipo: str, valor: str, color_fondo: str) -> str:
    """Aplica formato condicional a un rango.

    tipo: 'mayor_que' | 'menor_que' | 'igual_a' | 'contiene' | 'entre'
    valor: número/texto, o 'v1,v2' para 'entre'
    color_fondo: hex sin '#' (ej: 'FF0000' rojo)
    """
    try:
        from openpyxl.styles import PatternFill
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule

        wb = _cargar(ruta)
        ws = _hoja(wb, hoja)
        fill = PatternFill(fill_type="solid", fgColor=color_fondo.lstrip("#").upper())
        dxf = DifferentialStyle(fill=fill)

        if tipo == "entre":
            v1, v2 = [v.strip() for v in valor.split(",", 1)]
            rule = Rule(type="cellIs", operator="between", formula=[v1, v2], dxf=dxf)
        elif tipo == "contiene":
            rule = Rule(type="containsText", operator="containsText", text=valor, dxf=dxf)
        else:
            op_map = {"mayor_que": "greaterThan", "menor_que": "lessThan", "igual_a": "equal"}
            rule = Rule(type="cellIs", operator=op_map.get(tipo, "greaterThan"),
                        formula=[valor], dxf=dxf)

        ws.conditional_formatting.add(rango, rule)
        wb.save(ruta)
        return f"✅ Formato condicional '{tipo}' aplicado al rango {rango}"
    except Exception as e:
        logger.error(f"excel.formato_condicional: {e}")
        return f"Error: {e}"


# ── OBJETOS ───────────────────────────────────────────────────────────────────

def crear_tabla(ruta: str, hoja: str, rango: str, nombre: str,
                estilo: str = "TableStyleMedium9") -> str:
    """Crea una tabla Excel con estilo en el rango indicado."""
    try:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        wb = _cargar(ruta)
        ws = _hoja(wb, hoja)
        nombre_limpio = re.sub(r"[^A-Za-z0-9_]", "_", nombre)
        tabla = Table(displayName=nombre_limpio, ref=rango)
        tabla.tableStyleInfo = TableStyleInfo(
            name=estilo, showFirstColumn=False,
            showLastColumn=False, showRowStripes=True
        )
        ws.add_table(tabla)
        wb.save(ruta)
        return f"✅ Tabla '{nombre_limpio}' creada en {rango} con estilo '{estilo}'"
    except Exception as e:
        logger.error(f"excel.crear_tabla: {e}")
        return f"Error: {e}"


def agregar_grafico(ruta: str, hoja: str, tipo: str,
                    rango_datos: str, titulo: str = "",
                    celda_posicion: str = "E2") -> str:
    """Inserta un gráfico en la hoja.

    tipo: 'barras' | 'barras_apiladas' | 'lineas' | 'pastel' | 'area' | 'dispersion'
    rango_datos: ej. 'A1:B10'
    celda_posicion: celda donde se ancla el gráfico (ej. 'E2')
    """
    try:
        from openpyxl.chart import (BarChart, LineChart, PieChart,
                                    AreaChart, ScatterChart, Reference)
        wb = _cargar(ruta)
        ws = _hoja(wb, hoja)

        celdas = list(ws[rango_datos])
        min_row = celdas[0][0].row
        max_row = celdas[-1][0].row
        min_col = celdas[0][0].column
        max_col = celdas[0][-1].column

        data = Reference(ws, min_col=min_col, min_row=min_row,
                         max_col=max_col, max_row=max_row)

        tipo_map = {
            "barras": BarChart, "barras_apiladas": BarChart,
            "lineas": LineChart, "pastel": PieChart,
            "area": AreaChart, "dispersion": ScatterChart,
        }
        chart = tipo_map.get(tipo, BarChart)()
        chart.title = titulo
        if tipo == "barras_apiladas":
            chart.grouping = "stacked"
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, celda_posicion)
        wb.save(ruta)
        return f"✅ Gráfico '{tipo}' '{titulo}' insertado en {celda_posicion}"
    except Exception as e:
        logger.error(f"excel.agregar_grafico: {e}")
        return f"Error: {e}"


def proteger_hoja(ruta: str, hoja: str, contrasena: str = None) -> str:
    """Protege (con contraseña) o desprotege una hoja."""
    try:
        wb = _cargar(ruta)
        ws = _hoja(wb, hoja)
        if contrasena:
            ws.protection.sheet = True
            ws.protection.password = contrasena
            msg = f"✅ Hoja '{ws.title}' protegida"
        else:
            ws.protection.sheet = False
            msg = f"✅ Hoja '{ws.title}' desprotegida"
        wb.save(ruta)
        return msg
    except Exception as e:
        logger.error(f"excel.proteger_hoja: {e}")
        return f"Error: {e}"


def hipervinculo(ruta: str, hoja: str, celda: str, url: str, texto: str = None) -> str:
    """Agrega un hipervínculo en una celda."""
    try:
        from openpyxl.styles import Font
        wb = _cargar(ruta)
        ws = _hoja(wb, hoja)
        c = ws[celda]
        c.hyperlink = url
        c.value = texto or url
        c.font = Font(color="0563C1", underline="single")
        wb.save(ruta)
        return f"✅ Hipervínculo en {celda}: {url}"
    except Exception as e:
        logger.error(f"excel.hipervinculo: {e}")
        return f"Error: {e}"


# ── Dispatcher principal ──────────────────────────────────────────────────────

_OPERACIONES = {
    "info": info,
    "leer": leer,
    "leer_formulas": leer_formulas,
    "listar_hojas": listar_hojas,
    "buscar": buscar,
    "escribir": escribir,
    "crear": crear,
    "reemplazar": reemplazar,
    "crear_hoja": crear_hoja,
    "eliminar_hoja": eliminar_hoja,
    "renombrar_hoja": renombrar_hoja,
    "copiar_hoja": copiar_hoja,
    "mover_hoja": mover_hoja,
    "insertar_filas": insertar_filas,
    "eliminar_filas": eliminar_filas,
    "insertar_columnas": insertar_columnas,
    "eliminar_columnas": eliminar_columnas,
    "ordenar": ordenar,
    "formato_celdas": formato_celdas,
    "auto_ajustar_columnas": auto_ajustar_columnas,
    "formato_condicional": formato_condicional,
    "crear_tabla": crear_tabla,
    "agregar_grafico": agregar_grafico,
    "proteger_hoja": proteger_hoja,
    "hipervinculo": hipervinculo,
}


def ejecutar(operacion: str, **kwargs) -> str:
    """Punto de entrada único. Despacha a la función correspondiente."""
    fn = _OPERACIONES.get(operacion)
    if fn is None:
        disponibles = ", ".join(sorted(_OPERACIONES.keys()))
        return f"Operación '{operacion}' no existe. Disponibles: {disponibles}"
    try:
        return fn(**kwargs)
    except TypeError as e:
        return f"Parámetros incorrectos para '{operacion}': {e}"
    except Exception as e:
        logger.error(f"excel.ejecutar({operacion}): {e}", exc_info=True)
        return f"Error en '{operacion}': {e}"
